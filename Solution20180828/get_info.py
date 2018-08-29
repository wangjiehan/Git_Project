'''
用openpyxl方式在源文件上读取、修改并生成.xlsx
'''
import os
print(os.path.realpath(__file__))
print(os.path.dirname(os.path.realpath(__file__)))
print(os.path.basename(os.path.realpath(__file__)))
'''
读取本地当前文件夹下所有文件的路径
'''
def file_path(file_dir):
	for root, dirs, files in os.walk(file_dir):
		return root, files
a = file_path(os.path.dirname(os.path.realpath(__file__)))
print(a[0])
print(a[1])

from openpyxl import load_workbook
wb = load_workbook(filename = r'视频测试集.xlsx')
sheetnames = wb.sheetnames
print(sheetnames)
ws = wb[sheetnames[0]]
ws.cell(row = 1, column = 8).value = '本地路径'
wb.save(filename = '视频测试集.xlsx')
for i in range(len(a[1]) - 2):		# 当前文件夹中还有两个非视频文件
	ws.cell(row = 2 + i, column = 8).value = str(a[0] + '\\' + a[1][i])
wb.save('视频测试集.xlsx')

