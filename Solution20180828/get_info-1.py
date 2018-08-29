'''
获取本地视频文件路径，将数据存入一个新建的.xls/.xlsx文件
'''
import os
print(os.path.realpath(__file__))
print(os.path.dirname(os.path.realpath(__file__)))
print(os.path.basename(os.path.realpath(__file__)))

def file_path(file_dir):
	for root, dirs, files in os.walk(file_dir):
		return root, files
a = file_path(os.path.dirname(os.path.realpath(__file__)))

import xlwt
xls = xlwt.Workbook()
sht1 = xls.add_sheet('Local_address')
sht1.write(0, 0, '本地路径')
for i in range(len(a[1]) - 1):
	sht1.write(1 + i, 0, str(a[0] + '\\' + a[1][i]))
xls.save(r'C:\Task\Task20180828\Local_Address.xls')

'''
from openpyxl import Workbook
wb = Workbook()
ws = wb.worksheets[0]
ws.title = "Local_address"
ws.cell(row=1, column=1, value='本地路径')
for i in range(len(a[1]) - 2):
	ws.cell(1 + i, 1, str(a[0] + '\\' + a[1][i]))
wb.save(r'C:\Task\Task20180828\Local_Address.xlsx')
'''

